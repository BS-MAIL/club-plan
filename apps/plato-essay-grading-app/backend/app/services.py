import base64
import asyncio
import json
import re
import shutil
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

import fitz
import httpx
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from PIL import Image
from sqlalchemy.orm import Session

try:
    from google import genai
    from google.genai import types as genai_types
except ImportError:  # pragma: no cover
    genai = None
    genai_types = None

from app.config import get_settings
from app import models


settings = get_settings()

GRADING_JOBS: dict[str, dict] = {}
DB_WRITE_LOCK = threading.Lock()


def _is_sqlite_database() -> bool:
    return settings.database_url.startswith("sqlite")


def _sqlite_safe_concurrency(configured: int) -> int:
    """SQLite allows many readers but only one writer.

    Batch OCR/grading does a read -> slow AI call -> delete/insert/commit cycle for
    every item. Running several of those cycles at once often raises
    "database is locked" on Windows, so keep SQLite batch workers serial.
    """
    if _is_sqlite_database():
        return 1
    return max(1, configured)


VISION_PROMPT = """
This image is one cropped answer area from a Korean handwritten math constructed-response exam.
Do not grade the answer.

Tasks:
1. Transcribe visible Korean text as faithfully as possible.
2. Convert all mathematical expressions to LaTeX.
3. Summarize the student's solution steps in order.
4. Do not invent unclear symbols. Put uncertain or unreadable parts in uncertain_parts.
5. If the answer area is blank, has only meaningless marks, or has no meaningful answer, set korean_text to "답변 없음", normalized_solution to "답변 없음", confidence at least 0.9, and needs_review false.
6. If the answer includes a coordinate plane, graph, diagram, table, or marked points, describe it in normalized_solution. For graphs, identify the shape, opening/direction, vertex, axis, intercepts, plotted points, labels, and whether the student's drawn curve/line appears distinct from the printed grid or axes.
7. Return only valid JSON.

Required JSON schema:
{
  "korean_text": "string",
  "latex": ["string"],
  "normalized_solution": "string",
  "uncertain_parts": ["string"],
  "uncertain_marks": [
    {"text": "string", "reason": "string", "bbox": {"x": 0.0, "y": 0.0, "width": 0.0, "height": 0.0}, "confidence": 0.0}
  ],
  "confidence": 0.0,
  "needs_review": true
}

For uncertain_marks, use coordinates relative to the cropped answer image, from 0 to 1. Include only meaningful uncertainty that could affect grading.
"""

VISION_PROMPT_NO_FORMULA = """
This image is one cropped answer area from a Korean handwritten constructed-response exam.
Do not grade the answer.

Tasks:
1. Transcribe visible Korean text as faithfully as possible.
2. Summarize the student's answer and reasoning in order.
3. Do not convert expressions to LaTeX. Keep symbols and numbers in plain text as they appear.
4. Do not invent unclear words or symbols. Put uncertain or unreadable parts in uncertain_parts.
5. If the answer area is blank, has only meaningless marks, or has no meaningful answer, set korean_text to "답변 없음", normalized_solution to "답변 없음", confidence at least 0.9, and needs_review false.
6. If the answer includes a coordinate plane, graph, diagram, table, or marked points, describe it in normalized_solution. For graphs, identify the shape, opening/direction, vertex, axis, intercepts, plotted points, labels, and whether the student's drawn curve/line appears distinct from the printed grid or axes.
7. Return only valid JSON.

Required JSON schema:
{
  "korean_text": "string",
  "latex": [],
  "normalized_solution": "string",
  "uncertain_parts": ["string"],
  "uncertain_marks": [
    {"text": "string", "reason": "string", "bbox": {"x": 0.0, "y": 0.0, "width": 0.0, "height": 0.0}, "confidence": 0.0}
  ],
  "confidence": 0.0,
  "needs_review": true
}

For uncertain_marks, use coordinates relative to the cropped answer image, from 0 to 1. Include only meaningful uncertainty that could affect grading.
"""

ANSWER_TEMPLATE_DETECT_PROMPT = """
This image is a blank Korean math answer sheet template page.
Detect answer boxes or blank response regions for numbered questions.

Return only valid JSON:
{
  "regions": [
    {"question_no": 1, "x": 0, "y": 0, "width": 100, "height": 100, "confidence": 0.9}
  ]
}

Coordinate rules:
- x, y, width, height must be pixel coordinates in the provided image.
- Use the visible answer area, not the whole page.
- Match visible question numbers when possible.
- If a question number is unclear, omit that region.
- Prefer boxes/ruled response areas where students write answers.
"""

REFERENCE_IMPORT_PROMPT = """
You are reading Korean high school math constructed-response exam reference PDFs.
The upload may include a question document, an answer/rubric document, or both.

Extract and align them by question number. Return only valid JSON with this schema:
{
  "questions": [
    {
      "question_no": 1,
      "max_score": 4,
      "question_text": "original problem text, including conditions and math expressions in plain text where possible",
      "model_answer": "official answer, sample solution, or expected reasoning",
      "rubric_text": "only concise step scoring lines in the form '[단계별 채점 기준]\\n1. step name (points): criterion'. Use short Korean step names based on the question and model answer. Do not include partial-credit notes or general cautions."
    }
  ]
}

Rules:
- Use every uploaded document as reference. If both documents are present, align the original questions with the model answers and grading rubrics by question number.
- If only the question document is present, extract the original questions and create a reasonable minimal model_answer/rubric_text from the question content. Mark uncertain criteria with "[확인 필요]".
- If only the answer/rubric document is present, extract question numbers, max scores, model answers, and grading rubrics from that document. Leave question_text empty only when the original problem text is not visible.
- Use the visible question numbers. If a question number is unclear, infer only from neighboring numbering.
- If one main question has subquestions such as (1), (2), ㄱ, ㄴ, keep it as one item with the main question_no. Do not create question_no 71 or 72 for 7-(1), 7-(2).
- For subquestions, include the subquestion label in each rubric step name, for example: 1. (1) 판단 (1점): ... and 2. (2) 계산 (2점): ...
- If the answer/rubric PDF gives separate subquestion scores, sum them into max_score for the main question.
- Preserve Korean wording faithfully.
- Prefer plain-text math expressions for JSON stability.
- Do not use raw LaTeX delimiters such as \( \), \[ \], or unescaped backslash commands.
- If a backslash is unavoidable, escape it as a double backslash so the JSON remains valid.
- If max score is not visible, use 0.
- For rubric_text, create 2 to 6 scoring steps whose point total equals max_score when possible.
- Good rubric_text example: [단계별 채점 기준]\n1. 판단 (1점): 옳지 않음으로 판단\n2. 반례제시 (2점): 규칙적이지만 순환하지 않는 무한소수 제시\n3. 이유제시 (1점): 분수로 나타낼 수 없는 이유 제시
- Good subquestion example: [단계별 채점 기준]\n1. (1) 판단 (1점): 첫 번째 소문항의 판단 제시\n2. (1) 이유제시 (1점): 첫 번째 소문항의 근거 설명\n3. (2) 식 설정 (1점): 두 번째 소문항의 풀이식 설정\n4. (2) 결론 (1점): 두 번째 소문항의 답 도출
- Do not include sections named 유사풀이, 부분점 기준, 감점 기준, or 검토 필요 기준.
- If detailed scoring is unclear, do not leave rubric_text empty. Create a minimal rubric_text from the question and model answer, and add "[확인 필요]" to uncertain step criteria.
- Do not invent grading criteria that are not present.
- Return JSON only.
"""


def _escape_invalid_json_backslashes(content: str) -> str:
    return re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', content)


def _escape_latex_json_backslashes(content: str) -> str:
    return re.sub(r'(?<!\\)\\(?!(?:u[0-9a-fA-F]{4}))(?=(?:[A-Za-z]{2,}|[()\[\]{}]))', r'\\\\', content)


def _json_error_excerpt(content: str, error: json.JSONDecodeError) -> str:
    start = max(0, error.pos - 180)
    end = min(len(content), error.pos + 180)
    return content[start:end].replace("\n", " ")


def _loads_json_object(content: str) -> dict:
    latex_safe = _escape_latex_json_backslashes(content)
    if latex_safe != content:
        try:
            return json.loads(latex_safe)
        except json.JSONDecodeError:
            pass
    try:
        return json.loads(content)
    except json.JSONDecodeError as first_error:
        fixed = _escape_invalid_json_backslashes(content)
        if fixed != content:
            try:
                return json.loads(fixed)
            except json.JSONDecodeError:
                pass
        excerpt = _json_error_excerpt(content, first_error)
        raise ValueError(
            f"AI returned invalid JSON: {first_error.msg} at line {first_error.lineno}, column {first_error.colno}. "
            f"Nearby text: {excerpt}"
        ) from first_error


def _parse_json_object(content: str) -> dict:
    try:
        return _loads_json_object(content)
    except ValueError:
        start = content.find("{")
        end = content.rfind("}")
        if start >= 0 and end > start:
            return _loads_json_object(content[start:end + 1])
        raise


def _repair_json_object_with_openai(client, content: str) -> dict:
    response = client.chat.completions.create(
        model=settings.openai_model,
        messages=[{
            "role": "user",
            "content": (
                "다음 내용은 JSON으로 반환되어야 했지만 문자열 따옴표/쉼표/이스케이프 오류로 파싱에 실패했습니다. "
                "의미를 바꾸지 말고 유효한 JSON object 하나로만 고쳐서 반환하세요. 설명, markdown, 코드블록은 쓰지 마세요.\n\n"
                f"{content}"
            ),
        }],
        response_format={"type": "json_object"},
    )
    return _parse_json_object(response.choices[0].message.content or "{}")


def _parse_json_object_with_repair(client, content: str) -> dict:
    try:
        return _parse_json_object(content)
    except ValueError:
        if client is None or not settings.openai_api_key:
            raise
        return _repair_json_object_with_openai(client, content)


def _failed_vision_result(provider: str, error: Exception) -> dict:
    message = f"{type(error).__name__}: {error}"
    return _normalize_vision_result({
        "korean_text": "",
        "latex": [],
        "normalized_solution": "",
        "uncertain_parts": [f"{provider} OCR failed and needs manual review: {message}"],
        "uncertain_marks": [],
        "confidence": 0,
        "needs_review": True,
    }, provider)


def _failed_image_interpretation_result(error: Exception) -> dict:
    message = f"AI image interpretation failed: {type(error).__name__}: {error}"
    return {
        "text": "",
        "latex": "",
        "confidence": 0,
        "needs_review": True,
        "normalized": f"[AI image interpretation notes]\n{message}\nManual OCR review is required.",
        "uncertain_marks": [],
    }


def _normalize_vision_result(data: dict, provider: str) -> dict:
    latex = data.get("latex", [])
    if isinstance(latex, str):
        latex = [latex]
    uncertain_parts = data.get("uncertain_parts", [])
    if isinstance(uncertain_parts, str):
        uncertain_parts = [uncertain_parts]
    uncertain_marks = data.get("uncertain_marks", [])
    if not isinstance(uncertain_marks, list):
        uncertain_marks = []
    confidence = float(data.get("confidence", 0) or 0)
    needs_review = bool(data.get("needs_review", False)) or confidence < settings.ocr_confidence_threshold
    return {
        "provider": provider,
        "korean_text": str(data.get("korean_text", "")),
        "latex": [str(item) for item in latex],
        "normalized_solution": str(data.get("normalized_solution", "")),
        "uncertain_parts": [str(item) for item in uncertain_parts],
        "uncertain_marks": [item for item in uncertain_marks if isinstance(item, dict)],
        "confidence": confidence,
        "needs_review": needs_review,
    }


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value, default: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _review_reasons(question: models.Question, ocr: models.OcrResult, scores: list[float], reviews: list[dict]) -> list[str]:
    reasons = []
    max_deviation = max(scores) - min(scores) if scores else 0
    if ocr.needs_review and ocr.confidence < settings.ocr_confidence_threshold:
        reasons.append("ocr_review")
    if ocr.confidence < settings.ocr_confidence_threshold:
        reasons.append("ocr_low_confidence")
    if max_deviation >= max(2, question.max_score * 0.3):
        reasons.append("score_deviation")
    if not reviews or any(float(item.get("confidence", 0) or 0) < 0.3 for item in reviews):
        reasons.append("grading_uncertain")
    return list(dict.fromkeys(reasons))


NO_ANSWER_LABEL = "답변 없음"


def _is_no_answer_ocr(ocr: models.OcrResult) -> bool:
    text = "\n".join([ocr.text_result or "", ocr.latex_result or "", ocr.normalized_result or ""]).strip()
    if not text:
        return True
    lowered = text.lower()
    if "ai image interpretation failed" in lowered or "not configured" in lowered or "manual ocr review" in lowered:
        return False
    no_answer_patterns = [
        r"답변\s*없음",
        r"답안\s*없음",
        r"무응답",
        r"빈\s*칸",
        r"빈칸\s*제출",
        r"풀이\s*없음",
        r"의미\s*있는\s*답(?:변)?\s*(?:이\s*)?없",
        r"blank",
        r"no\s+answer",
        r"no\s+meaningful\s+answer",
    ]
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in no_answer_patterns)


def _save_no_answer_grade(db: Session, submission_id: int, question_id: int) -> dict:
    reason = "답안이 비어 있거나 유의미한 답변이 없어 0점 처리했습니다."
    with DB_WRITE_LOCK:
        db.query(models.GradingReview).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission_id, question_id=question_id, annotation_type="deduction").delete()
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission_id, question_id=question_id, annotation_type="blank").delete()
        db.query(models.FinalScore).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.add(models.GradingReview(
            submission_id=submission_id,
            question_id=question_id,
            round_no=1,
            score=0,
            confidence=1,
            reasoning=NO_ANSWER_LABEL,
            deduction_reasons=NO_ANSWER_LABEL,
            missing_steps="유의미한 답변 없음",
            ocr_risk="low",
            similar_answer_review=False,
            similar_answer_reason="",
        ))
        db.add(models.FinalScore(
            submission_id=submission_id,
            question_id=question_id,
            final_score=0,
            final_reason=NO_ANSWER_LABEL,
            needs_teacher_review=False,
        ))
        _add_annotation(db, submission_id, question_id, "blank", {
            "label": NO_ANSWER_LABEL,
            "reason": reason,
            "confidence": 1,
        }, NO_ANSWER_LABEL)
        db.commit()
    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "final_score": 0,
        "needs_teacher_review": False,
        "review_scores": [0],
        "max_deviation": 0,
        "final_reason": NO_ANSWER_LABEL,
        "review_reasons": [],
        "similar_answer_review": False,
        "similar_answer_reason": "",
    }


def _fallback_grade_for_failed_ai(db: Session, submission_id: int, question_id: int, reason: str) -> dict:
    with DB_WRITE_LOCK:
        db.query(models.GradingReview).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission_id, question_id=question_id, annotation_type="deduction").delete()
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission_id, question_id=question_id, annotation_type="blank").delete()
        db.query(models.FinalScore).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.add(models.GradingReview(
            submission_id=submission_id,
            question_id=question_id,
            round_no=1,
            score=0,
            confidence=0,
            reasoning=reason,
            deduction_reasons="Manual grading required because AI processing failed.",
            missing_steps="AI processing failed; the answer must be checked manually.",
            ocr_risk="high",
            similar_answer_review=False,
            similar_answer_reason="",
        ))
        db.add(models.FinalScore(
            submission_id=submission_id,
            question_id=question_id,
            final_score=0,
            final_reason=reason,
            needs_teacher_review=True,
        ))
        _add_annotation(db, submission_id, question_id, "blank", {
            "label": "AI 처리 실패",
            "reason": reason,
            "confidence": 1,
        }, "AI 처리 실패")
        db.commit()
    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "final_score": 0,
        "needs_teacher_review": True,
        "review_scores": [0],
        "max_deviation": 0,
        "final_reason": reason,
        "review_reasons": ["ocr_review", "ocr_low_confidence", "grading_uncertain"],
        "similar_answer_review": False,
        "similar_answer_reason": "",
    }


def _fallback_grade_for_failed_ocr(db: Session, submission_id: int, question_id: int) -> dict:
    return _fallback_grade_for_failed_ai(
        db,
        submission_id,
        question_id,
        "OCR/이미지 해석에 실패하여 자동 채점 대신 교사 확인이 필요합니다.",
    )


def _parse_deductions(value: str) -> list[dict]:
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    deductions = []
    for item in parsed:
        if not isinstance(item, dict):
            continue
        deductions.append({
            "step_name": str(item.get("step_name", "감점")),
            "points_lost": _safe_float(item.get("points_lost"), 0),
            "reason": str(item.get("reason", "")),
            "short_reason": str(item.get("short_reason", "")),
            "ocr_related": bool(item.get("ocr_related", False)),
        })
    return deductions


def _bbox_from_item(item: dict) -> dict | None:
    bbox = item.get("bbox") if isinstance(item, dict) else None
    if not isinstance(bbox, dict):
        return None
    x = _safe_float(bbox.get("x"), -1)
    y = _safe_float(bbox.get("y"), -1)
    width = _safe_float(bbox.get("width"), 0)
    height = _safe_float(bbox.get("height"), 0)
    if x < 0 or y < 0 or width <= 0 or height <= 0:
        return None
    return {
        "x": max(0, min(1, x)),
        "y": max(0, min(1, y)),
        "width": max(0.01, min(1, width)),
        "height": max(0.01, min(1, height)),
    }


def _add_annotation(db: Session, submission_id: int, question_id: int, annotation_type: str, item: dict, fallback_label: str = "") -> None:
    bbox = _bbox_from_item(item)
    db.add(models.AnswerAnnotation(
        submission_id=submission_id,
        question_id=question_id,
        annotation_type=annotation_type,
        label=str(item.get("label") or fallback_label),
        reason=str(item.get("reason", "")),
        evidence_text=str(item.get("evidence_text") or item.get("text") or ""),
        x=bbox["x"] if bbox else None,
        y=bbox["y"] if bbox else None,
        width=bbox["width"] if bbox else None,
        height=bbox["height"] if bbox else None,
        confidence=_safe_float(item.get("confidence"), 0),
    ))


def _has_step_rubric(text: str) -> bool:
    return bool(re.search(r"^\s*\d+\.\s*.+?(?:\(\d+(?:\.\d+)?점\))?\s*[:：]", text or "", re.MULTILINE))


def _fallback_rubric_text(question_no: int, max_score: float, question_text: str, model_answer: str) -> str:
    sub_labels = re.findall(r"\((\d+)\)", f"{question_text}\n{model_answer}")
    labels = list(dict.fromkeys(sub_labels))
    if labels:
        per_step = round(max_score / len(labels), 1) if max_score and labels else 0
        lines = ["[단계별 채점 기준]"]
        for index, label in enumerate(labels, start=1):
            lines.append(f"{index}. ({label}) 풀이 ({per_step:g}점): 소문항 ({label})의 핵심 판단, 계산, 결론을 제시 [확인 필요]")
        return "\n".join(lines)

    if max_score > 0:
        first = round(max_score * 0.4, 1)
        second = round(max_score - first, 1)
    else:
        first = 0
        second = 0
    return "\n".join([
        "[단계별 채점 기준]",
        f"1. 핵심판단 ({first:g}점): 문제 조건에 맞는 핵심 판단 또는 식을 제시 [확인 필요]",
        f"2. 풀이완성 ({second:g}점): 모범답안에 맞는 계산, 이유, 결론을 완성 [확인 필요]",
    ])


def get_openai_client() -> OpenAI:
    kwargs = {"api_key": settings.openai_api_key}
    if settings.openai_base_url:
        kwargs["base_url"] = settings.openai_base_url.rstrip("/")
    return OpenAI(**kwargs)


def save_upload(file, directory: str, filename: str) -> Path:
    target_dir = Path(directory)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / filename
    with target.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return target


def import_students_from_excel(db: Session, exam_id: int, file_path: Path) -> int:
    clear_exam_students_and_submissions(db, exam_id)
    workbook = load_workbook(file_path)
    sheet = workbook.active
    count = 0
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if not row or row[0] is None or row[1] is None:
            continue
        absent_value = str(row[2]).strip().lower() if len(row) > 2 and row[2] is not None else ""
        is_absent = absent_value in {"미응시", "결석", "absent", "x", "n", "no", "0"}
        db.add(models.Student(exam_id=exam_id, student_no=str(row[0]), name=str(row[1]), order_index=index, is_absent=is_absent))
        count += 1
    db.commit()
    return count


class PdfPageCountMismatch(ValueError):
    def __init__(self, actual_pages: int, expected_pages: int, student_count: int, pages_per_student: int, ignored_pages: list[int] | None = None, absent_count: int = 0):
        self.actual_pages = actual_pages
        self.expected_pages = expected_pages
        self.student_count = student_count
        self.absent_count = absent_count
        self.pages_per_student = pages_per_student
        self.ignored_pages = ignored_pages or []
        remaining_pages = actual_pages - len(self.ignored_pages)
        self.remaining_pages = remaining_pages
        super().__init__(f"PDF page count {remaining_pages} does not match expected {expected_pages} pages")

    def to_detail(self) -> dict:
        return {
            "code": "pdf_page_count_mismatch",
            "message": str(self),
            "actual_pages": self.actual_pages,
            "expected_pages": self.expected_pages,
            "remaining_pages": self.remaining_pages,
            "difference": self.remaining_pages - self.expected_pages,
            "student_count": self.student_count,
            "absent_count": self.absent_count,
            "present_count": self.student_count - self.absent_count,
            "pages_per_student": self.pages_per_student,
            "ignored_pages": self.ignored_pages,
        }


def parse_ignored_pages(value: str | None, page_count: int) -> list[int]:
    if not value or not value.strip():
        return []
    ignored: set[int] = set()
    for raw_part in value.replace(" ", "").split(","):
        if not raw_part:
            continue
        if "-" in raw_part:
            start_text, end_text = raw_part.split("-", 1)
            start = int(start_text)
            end = int(end_text)
            if start > end:
                start, end = end, start
            for page_no in range(start, end + 1):
                if page_no < 1 or page_no > page_count:
                    raise ValueError(f"Ignored page {page_no} is out of range 1-{page_count}")
                ignored.add(page_no)
        else:
            page_no = int(raw_part)
            if page_no < 1 or page_no > page_count:
                raise ValueError(f"Ignored page {page_no} is out of range 1-{page_count}")
            ignored.add(page_no)
    return sorted(ignored)


def split_pdf_into_submissions(db: Session, exam: models.Exam, pdf_path: Path, ignored_pages_text: str | None = None) -> int:
    clear_exam_submissions(db, exam.id)
    document = fitz.open(pdf_path)
    actual_pages = document.page_count
    all_students = db.query(models.Student).filter_by(exam_id=exam.id).order_by(models.Student.order_index).all()
    students = [student for student in all_students if not student.is_absent]
    student_count = len(all_students)
    absent_count = student_count - len(students)
    expected_pages = len(students) * exam.pages_per_student
    ignored_pages = parse_ignored_pages(ignored_pages_text, actual_pages)
    usable_pages = [page_index for page_index in range(actual_pages) if page_index + 1 not in set(ignored_pages)]
    if len(usable_pages) != expected_pages:
        document.close()
        raise PdfPageCountMismatch(actual_pages, expected_pages, student_count, exam.pages_per_student, ignored_pages, absent_count)

    output_dir = Path(settings.processed_dir) / f"exam_{exam.id}" / "submissions"
    output_dir.mkdir(parents=True, exist_ok=True)
    for idx, student in enumerate(students):
        start = idx * exam.pages_per_student
        end = start + exam.pages_per_student
        student_pages = usable_pages[start:end]
        student_doc = fitz.open()
        for page_index in student_pages:
            student_doc.insert_pdf(document, from_page=page_index, to_page=page_index)
        out_file = output_dir / f"{student.order_index:03d}_{student.student_no}.pdf"
        student_doc.save(out_file)
        student_doc.close()
        db.add(models.Submission(
            exam_id=exam.id,
            student_id=student.id,
            start_page=student_pages[0] + 1,
            end_page=student_pages[-1] + 1,
            file_path=str(out_file),
            match_status="order_matched",
        ))
    db.commit()
    document.close()
    return len(students)


def update_student_absences(db: Session, exam_id: int, absent_student_ids: list[int]) -> list[models.Student]:
    students = db.query(models.Student).filter_by(exam_id=exam_id).order_by(models.Student.order_index).all()
    if not students:
        return []
    valid_ids = {student.id for student in students}
    absent_ids = {student_id for student_id in absent_student_ids if student_id in valid_ids}
    for student in students:
        student.is_absent = student.id in absent_ids
    db.commit()
    return students


def crop_region_from_submission(submission: models.Submission, region: models.Region, output_path: Path) -> Path:
    doc = fitz.open(submission.file_path)
    page = doc[region.page_offset]
    page_width = page.rect.width
    page_height = page.rect.height
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    page_image = output_path.with_suffix(".page.png")
    pix.save(page_image)
    doc.close()

    image = Image.open(page_image)
    scale_x = image.width / page_width
    scale_y = image.height / page_height
    box = (
        int(region.x * scale_x),
        int(region.y * scale_y),
        int((region.x + region.width) * scale_x),
        int((region.y + region.height) * scale_y),
    )
    cropped = image.crop(box)
    cropped.save(output_path)
    page_image.unlink(missing_ok=True)
    return output_path


def render_submission_page_preview(submission: models.Submission, page_offset: int) -> Path:
    doc = fitz.open(submission.file_path)
    if page_offset < 0 or page_offset >= doc.page_count:
        doc.close()
        raise ValueError("Page offset is out of range")
    page = doc[page_offset]
    output_dir = Path(settings.processed_dir) / f"exam_{submission.exam_id}" / "previews" / f"submission_{submission.id}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"page_{page_offset}.png"
    if not output_path.exists():
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        pix.save(output_path)
    doc.close()
    return output_path


def render_submission_region_preview(submission: models.Submission, region: models.Region) -> Path:
    output_dir = Path(settings.processed_dir) / f"exam_{submission.exam_id}" / "region_previews" / f"submission_{submission.id}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"question_{region.question_id}.png"
    if not output_path.exists():
        crop_region_from_submission(submission, region, output_path)
    return output_path


def render_pdf_page_preview(pdf_path: Path, output_dir: Path, page_offset: int) -> Path:
    doc = fitz.open(pdf_path)
    if page_offset < 0 or page_offset >= doc.page_count:
        doc.close()
        raise ValueError("Page offset is out of range")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"page_{page_offset}.png"
    if not output_path.exists():
        pix = doc[page_offset].get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        pix.save(output_path)
    doc.close()
    return output_path


def get_answer_template_path(exam_id: int) -> Path:
    return Path(settings.upload_dir) / f"exam_{exam_id}_answer_template.pdf"


def save_answer_template(file, exam_id: int) -> Path:
    return save_upload(file, settings.upload_dir, f"exam_{exam_id}_answer_template.pdf")


def render_answer_template_preview(exam_id: int, page_offset: int) -> Path:
    path = get_answer_template_path(exam_id)
    if not path.exists():
        raise ValueError("Answer template PDF not found")
    return render_pdf_page_preview(path, Path(settings.processed_dir) / f"exam_{exam_id}" / "template_previews", page_offset)


async def run_mathpix(image_path: Path) -> dict:
    if not settings.mathpix_app_id or not settings.mathpix_app_key:
        return {
            "text": "Mathpix API key is not configured. Enter corrected OCR text manually.",
            "latex": "",
            "confidence": 0,
            "needs_review": True,
        }

    image_b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    payload = {
        "src": f"data:image/png;base64,{image_b64}",
        "formats": ["text", "latex_styled"],
        "ocr": ["math", "text"],
    }
    headers = {"app_id": settings.mathpix_app_id, "app_key": settings.mathpix_app_key}
    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post("https://api.mathpix.com/v3/text", json=payload, headers=headers)
        response.raise_for_status()
        data = response.json()
    return {
        "text": data.get("text", ""),
        "latex": data.get("latex_styled", ""),
        "confidence": float(data.get("confidence", 0) or 0),
        "needs_review": float(data.get("confidence", 0) or 0) < 0.75,
    }


async def run_openai_vision(image_path: Path, formula_recognition_enabled: bool = True) -> dict:
    if not settings.openai_api_key:
        return _normalize_vision_result({
            "korean_text": "OpenAI API key is not configured.",
            "latex": [],
            "normalized_solution": "",
            "uncertain_parts": ["OpenAI vision was skipped because OPENAI_API_KEY is missing."],
            "confidence": 0,
            "needs_review": True,
        }, "openai")

    image_b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    prompt = VISION_PROMPT if formula_recognition_enabled else VISION_PROMPT_NO_FORMULA
    client = get_openai_client()
    response = await asyncio.to_thread(
        client.chat.completions.create,
        model=settings.openai_model,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
            ],
        }],
        response_format={"type": "json_object"},
    )
    content = response.choices[0].message.content or "{}"
    try:
        parsed = _parse_json_object_with_repair(client, content)
    except Exception as exc:  # noqa: BLE001
        return _failed_vision_result("openai", exc)
    return _normalize_vision_result(parsed, "openai")


def _pdf_pages_as_data_urls(pdf_path: Path, max_pages: int = 20) -> list[str]:
    document = fitz.open(pdf_path)
    data_urls = []
    try:
        for page_index in range(min(document.page_count, max_pages)):
            page = document[page_index]
            pix = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
            image_b64 = base64.b64encode(pix.tobytes("png")).decode("ascii")
            data_urls.append(f"data:image/png;base64,{image_b64}")
    finally:
        document.close()
    return data_urls


async def import_reference_pdfs(db: Session, exam_id: int, question_pdf: Path | None, answer_pdf: Path | None) -> list[models.Question]:
    if not settings.openai_api_key:
        raise ValueError("OPENAI_API_KEY is required to import question and reference PDFs")

    question_pages = _pdf_pages_as_data_urls(question_pdf) if question_pdf else []
    answer_pages = _pdf_pages_as_data_urls(answer_pdf) if answer_pdf else []
    if not question_pages and not answer_pages:
        raise ValueError("At least one question or answer/rubric PDF must contain a page")

    content = [{"type": "text", "text": REFERENCE_IMPORT_PROMPT}]
    if question_pages:
        content.append({"type": "text", "text": "[문제 파일 시작]"})
        content.extend({"type": "image_url", "image_url": {"url": url}} for url in question_pages)
    if answer_pages:
        content.append({"type": "text", "text": "[모범답안 및 채점기준 파일 시작]"})
        content.extend({"type": "image_url", "image_url": {"url": url}} for url in answer_pages)

    client = get_openai_client()
    response = client.chat.completions.create(
        model=settings.openai_model,
        messages=[{"role": "user", "content": content}],
        response_format={"type": "json_object"},
    )
    parsed = _parse_json_object(response.choices[0].message.content or "{}")
    question_items = parsed.get("questions", [])
    if not isinstance(question_items, list):
        raise ValueError("AI response did not include a questions array")
    if not question_items:
        keys = ", ".join(parsed.keys()) if isinstance(parsed, dict) else "unknown"
        raise ValueError(
            "GPT did not find any questions in the uploaded PDFs. "
            "Check that the PDFs are readable, upright, and include visible question numbers. "
            f"AI response keys: {keys}"
        )

    imported = []
    for item in question_items:
        if not isinstance(item, dict):
            continue
        question_no = _safe_int(item.get("question_no"))
        if question_no <= 0:
            continue
        question = db.query(models.Question).filter_by(exam_id=exam_id, question_no=question_no).first()
        if not question:
            question = models.Question(exam_id=exam_id, question_no=question_no, max_score=0)
            db.add(question)
        question.max_score = _safe_float(item.get("max_score"), question.max_score or 0)
        question.question_text = str(item.get("question_text", "")).strip()
        question.model_answer = str(item.get("model_answer", "")).strip()
        rubric_text = str(item.get("rubric_text", "")).strip()
        if not _has_step_rubric(rubric_text):
            rubric_text = _fallback_rubric_text(question_no, question.max_score, question.question_text, question.model_answer)
        question.rubric_text = rubric_text
        imported.append(question)

    if not imported:
        raise ValueError(
            "GPT returned question-like data, but no valid positive question_no values were found. "
            "Check the PDF question numbering or try a clearer PDF export."
        )

    db.commit()
    for question in imported:
        db.refresh(question)
    return imported


async def run_gemini_vision(image_path: Path, formula_recognition_enabled: bool = True) -> dict:
    if not settings.gemini_api_key:
        return _normalize_vision_result({
            "korean_text": "Gemini API key is not configured.",
            "latex": [],
            "normalized_solution": "",
            "uncertain_parts": ["Gemini vision was skipped because GEMINI_API_KEY is missing."],
            "confidence": 0,
            "needs_review": True,
        }, "gemini")
    if genai is None or genai_types is None:
        return _normalize_vision_result({
            "korean_text": "google-genai package is not installed.",
            "latex": [],
            "normalized_solution": "",
            "uncertain_parts": ["Install google-genai to use Gemini vision."],
            "confidence": 0,
            "needs_review": True,
        }, "gemini")

    prompt = VISION_PROMPT if formula_recognition_enabled else VISION_PROMPT_NO_FORMULA
    client = genai.Client(api_key=settings.gemini_api_key)
    response = client.models.generate_content(
        model=settings.gemini_model,
        contents=[
            prompt,
            genai_types.Part.from_bytes(data=image_path.read_bytes(), mime_type="image/png"),
        ],
        config={"response_mime_type": "application/json"},
    )
    try:
        parsed = _parse_json_object_with_repair(get_openai_client() if settings.openai_api_key else None, response.text or "{}")
    except Exception as exc:  # noqa: BLE001
        return _failed_vision_result("gemini", exc)
    return _normalize_vision_result(parsed, "gemini")


def reconcile_vision_results(primary: dict, secondary: dict | None = None) -> dict:
    chosen = primary
    notes = []
    providers = [primary.get("provider", "openai")]
    if secondary:
        providers.append(secondary.get("provider", "gemini"))
        notes.append("Gemini was used because the primary OpenAI result was uncertain or low confidence.")
        if secondary["confidence"] > primary["confidence"] and secondary["normalized_solution"]:
            chosen = secondary
            notes.append("Gemini result had higher confidence and was selected as the main transcription.")
        else:
            notes.append("OpenAI result was kept as the main transcription.")

    uncertainty = list(chosen.get("uncertain_parts", []))
    uncertain_marks = list(chosen.get("uncertain_marks", []))
    if secondary:
        uncertainty.extend(secondary.get("uncertain_parts", []))
        uncertain_marks.extend(secondary.get("uncertain_marks", []))
    confidence = float(chosen.get("confidence", 0) or 0)
    severe_uncertainty = any(
        re.search(r"(unreadable|illegible|cannot read|cropped|잘림|판독 불가|읽을 수|확인 필요)", str(item), re.IGNORECASE)
        for item in uncertainty
    )
    needs_review = confidence < settings.ocr_confidence_threshold or severe_uncertainty
    latex_text = "\n".join(chosen.get("latex", []))
    text_result = chosen.get("korean_text", "")
    normalized = chosen.get("normalized_solution", "") or text_result
    if uncertainty or notes:
        normalized += "\n\n[AI image interpretation notes]\n" + json.dumps({
            "providers": providers,
            "notes": notes,
            "uncertain_parts": uncertainty,
            "uncertain_marks": uncertain_marks,
        }, ensure_ascii=False)
    return {
        "text": text_result,
        "latex": latex_text,
        "confidence": confidence,
        "needs_review": needs_review,
        "normalized": normalized,
        "uncertain_marks": uncertain_marks,
    }


async def run_ai_image_interpretation(image_path: Path, formula_recognition_enabled: bool = True) -> dict:
    try:
        primary = await run_openai_vision(image_path, formula_recognition_enabled=formula_recognition_enabled)
        secondary = None
        if primary["needs_review"] and settings.gemini_api_key:
            secondary = await run_gemini_vision(image_path, formula_recognition_enabled=formula_recognition_enabled)
        if formula_recognition_enabled and primary["needs_review"] and not settings.gemini_api_key and settings.mathpix_app_id and settings.mathpix_app_key:
            try:
                mathpix = await run_mathpix(image_path)
                return {
                    "text": mathpix["text"],
                    "latex": mathpix["latex"],
                    "confidence": mathpix["confidence"],
                    "needs_review": mathpix["needs_review"],
                    "normalized": mathpix["text"] + "\n\n[AI image interpretation notes]\nMathpix fallback was used.",
                    "uncertain_marks": [],
                }
            except Exception as exc:  # noqa: BLE001
                primary["uncertain_parts"].append(f"Mathpix fallback failed: {type(exc).__name__}: {exc}")
        return reconcile_vision_results(primary, secondary)
    except Exception as exc:  # noqa: BLE001
        return _failed_image_interpretation_result(exc)


def _expand_detected_region(region: dict, image_width: float, image_height: float) -> dict:
    x = _safe_float(region.get("x"), 0)
    y = _safe_float(region.get("y"), 0)
    width = _safe_float(region.get("width"), 0)
    height = _safe_float(region.get("height"), 0)
    margin_x = width * settings.region_margin_x_ratio
    margin_top = height * settings.region_margin_top_ratio
    margin_bottom = height * settings.region_margin_bottom_ratio
    expanded_x = max(0, x - margin_x)
    expanded_y = max(0, y - margin_top)
    expanded_right = min(image_width, x + width + margin_x)
    expanded_bottom = min(image_height, y + height + margin_bottom)
    return {
        "x": expanded_x / 2,
        "y": expanded_y / 2,
        "width": max(1, expanded_right - expanded_x) / 2,
        "height": max(1, expanded_bottom - expanded_y) / 2,
        "detected_x": x / 2,
        "detected_y": y / 2,
        "detected_width": width / 2,
        "detected_height": height / 2,
    }


async def detect_answer_template_regions(db: Session, exam_id: int) -> list[dict]:
    template_path = get_answer_template_path(exam_id)
    if not template_path.exists():
        raise ValueError("Answer template PDF not found")
    if not settings.openai_api_key:
        raise ValueError("OPENAI_API_KEY is required for automatic answer-region detection")

    questions = db.query(models.Question).filter_by(exam_id=exam_id).order_by(models.Question.question_no).all()
    question_by_no = {question.question_no: question for question in questions}
    doc = fitz.open(template_path)
    detected: list[dict] = []
    client = get_openai_client()
    try:
        for page_offset in range(doc.page_count):
            page = doc[page_offset]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image_b64 = base64.b64encode(pix.tobytes("png")).decode("ascii")
            response = await asyncio.to_thread(
                client.chat.completions.create,
                model=settings.openai_model,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": ANSWER_TEMPLATE_DETECT_PROMPT},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
                    ],
                }],
                response_format={"type": "json_object"},
            )
            parsed = _parse_json_object(response.choices[0].message.content or "{}")
            for item in parsed.get("regions", []):
                if not isinstance(item, dict):
                    continue
                question_no = _safe_int(item.get("question_no"), 0)
                question = question_by_no.get(question_no)
                if not question:
                    continue
                expanded = _expand_detected_region(item, pix.width, pix.height)
                detected.append({
                    "question_id": question.id,
                    "question_no": question_no,
                    "page_offset": page_offset,
                    **expanded,
                    "confidence": _safe_float(item.get("confidence"), 0),
                })
    finally:
        doc.close()
    return detected


async def create_ocr_for_submission(db: Session, submission_id: int, force: bool = False) -> int:
    submission = db.get(models.Submission, submission_id)
    if not submission:
        raise ValueError("Submission not found")
    exam = db.get(models.Exam, submission.exam_id)
    formula_recognition_enabled = exam.formula_recognition_enabled if exam else True
    regions = db.query(models.Region).filter_by(exam_id=submission.exam_id, region_type="question").all()
    submission_id = submission.id
    output_dir = Path(settings.processed_dir) / f"exam_{submission.exam_id}" / "ocr" / f"submission_{submission.id}"
    output_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for region in regions:
        if not region.question_id:
            continue
        question_id = region.question_id
        ocr = db.query(models.OcrResult).filter_by(submission_id=submission_id, question_id=question_id).first()
        if ocr and ocr.image_path and Path(ocr.image_path).exists() and not force:
            count += 1
            continue
        image_path = output_dir / f"question_{question_id}.png"
        crop_region_from_submission(submission, region, image_path)
        db.commit()
        try:
            result = await run_ai_image_interpretation(image_path, formula_recognition_enabled=formula_recognition_enabled)
        except Exception as exc:  # noqa: BLE001
            result = _failed_image_interpretation_result(exc)
        with DB_WRITE_LOCK:
            ocr = db.query(models.OcrResult).filter_by(submission_id=submission_id, question_id=question_id).first()
            if not ocr:
                ocr = models.OcrResult(submission_id=submission_id, question_id=question_id)
                db.add(ocr)
            ocr.image_path = str(image_path)
            ocr.text_result = result["text"]
            ocr.latex_result = result["latex"]
            ocr.normalized_result = result["normalized"]
            ocr.confidence = result["confidence"]
            ocr.needs_review = result["needs_review"]
            db.query(models.AnswerAnnotation).filter_by(
                submission_id=submission_id,
                question_id=question_id,
                annotation_type="ocr",
            ).delete()
            for mark in result.get("uncertain_marks", []):
                if isinstance(mark, dict):
                    _add_annotation(db, submission_id, question_id, "ocr", mark, "OCR 확인")
            db.commit()
        count += 1
    db.expire_all()
    return count


async def create_ocr_for_submission_question(db: Session, submission_id: int, question_id: int, force: bool = False) -> int:
    submission = db.get(models.Submission, submission_id)
    if not submission:
        raise ValueError("Submission not found")
    region = db.query(models.Region).filter_by(exam_id=submission.exam_id, region_type="question", question_id=question_id).first()
    if not region:
        raise ValueError("Question region not found")
    output_dir = Path(settings.processed_dir) / f"exam_{submission.exam_id}" / "ocr" / f"submission_{submission.id}"
    output_dir.mkdir(parents=True, exist_ok=True)
    ocr = db.query(models.OcrResult).filter_by(submission_id=submission.id, question_id=question_id).first()
    if ocr and ocr.image_path and Path(ocr.image_path).exists() and not force:
        return 1
    image_path = output_dir / f"question_{question_id}.png"
    crop_region_from_submission(submission, region, image_path)
    db.commit()
    try:
        exam = db.get(models.Exam, submission.exam_id)
        result = await run_ai_image_interpretation(image_path, formula_recognition_enabled=exam.formula_recognition_enabled if exam else True)
    except Exception as exc:  # noqa: BLE001
        result = _failed_image_interpretation_result(exc)
    with DB_WRITE_LOCK:
        ocr = db.query(models.OcrResult).filter_by(submission_id=submission.id, question_id=question_id).first()
        if not ocr:
            ocr = models.OcrResult(submission_id=submission.id, question_id=question_id)
            db.add(ocr)
        ocr.image_path = str(image_path)
        ocr.text_result = result["text"]
        ocr.latex_result = result["latex"]
        ocr.normalized_result = result["normalized"]
        ocr.confidence = result["confidence"]
        ocr.needs_review = result["needs_review"]
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission.id, question_id=question_id, annotation_type="ocr").delete()
        for mark in result.get("uncertain_marks", []):
            if isinstance(mark, dict):
                _add_annotation(db, submission.id, question_id, "ocr", mark, "OCR 확인")
        db.commit()
    return 1


def _fallback_reviews(question: models.Question, ocr: models.OcrResult) -> list[dict]:
    needs_review = ocr.needs_review or not ocr.text_result.strip() or "not configured" in ocr.text_result
    score = 0 if needs_review else round(question.max_score * 0.5, 1)
    return [
        {
            "round_no": i,
            "score": score,
            "confidence": 0.1,
            "reasoning": "OpenAI API key is not configured, so this is a placeholder review.",
            "deduction_reasons": "Manual grading required.",
            "missing_steps": "Unknown until AI grading or teacher review is performed.",
            "ocr_risk": "high" if needs_review else "medium",
            "deductions": [],
        }
        for i in range(1, 5)
    ]


async def grade_submission_question(db: Session, submission_id: int, question_id: int, fast: bool = True) -> dict:
    question = db.get(models.Question, question_id)
    ocr = db.query(models.OcrResult).filter_by(submission_id=submission_id, question_id=question_id).first()
    if not question or not ocr:
        raise ValueError("Question or OCR result not found")
    max_score = question.max_score
    question_text = question.question_text
    model_answer = question.model_answer
    rubric_text = question.rubric_text
    submission = db.get(models.Submission, submission_id)
    exam = db.get(models.Exam, submission.exam_id) if submission else None
    formula_recognition_enabled = exam.formula_recognition_enabled if exam else True
    ocr_text = ocr.text_result
    ocr_latex = ocr.latex_result
    ocr_normalized = ocr.normalized_result
    ocr_needs_review = ocr.needs_review
    ocr_confidence = ocr.confidence
    grading_subject = "Korean handwritten math constructed-response answer" if formula_recognition_enabled else "Korean handwritten constructed-response answer"
    similar_answers = db.query(models.SimilarAnswer).filter_by(question_id=question_id).order_by(models.SimilarAnswer.id).all()
    similar_answer_text = "\n".join(
        f"- {item.title or '유사답안'}: {item.text}" for item in similar_answers if item.text.strip()
    )
    db.commit()

    if _is_no_answer_ocr(ocr):
        return _save_no_answer_grade(db, submission_id, question_id)

    if not settings.openai_api_key:
        reviews = _fallback_reviews(question, ocr)
    else:
        client = get_openai_client()
        review_instruction = (
            "Return only valid JSON with key reviews, an array with one review object. "
            "Use this fast pass only when the answer is clear. If uncertain, lower confidence below 0.75. "
            "The review object must include round_no, score, confidence, reasoning, deduction_reasons, missing_steps, ocr_risk, deductions, similar_answer_review, similar_answer_reason."
        ) if fast else (
            "Return only valid JSON with key reviews, an array of four review objects. "
            "Each review object must include round_no, score, confidence, reasoning, deduction_reasons, missing_steps, ocr_risk, deductions, similar_answer_review, similar_answer_reason."
        )
        round_roles = "Fast pass: one rubric-based grading review." if fast else "Round roles: 1 strict rubric grading, 2 independent grading, 3 OCR/math error audit, 4 final reconciliation."
        prompt = f"""
You are grading a {grading_subject}.
{review_instruction}
deductions must be an array of objects: step_name, points_lost, reason, short_reason, ocr_related.
For each deduction, short_reason must be a concise Korean summary for teachers, preferably 10-25 Korean characters, focusing on the missing condition/value or exact error. Examples: "b=-16 누락", "a 부호 조건 누락", "계산식 전개 오류", "근거 설명 부족", "범위 조건 누락". Do not copy the full reason into short_reason.
Each deduction should also include evidence_text and bbox when the wrong expression/phrase can be located in the answer image.
bbox coordinates must be relative to the cropped answer image, from 0 to 1: {{"x": 0.0, "y": 0.0, "width": 0.0, "height": 0.0}}.
Example: if the wrong written expression is x=2, set evidence_text to "x=2" and bbox around that expression.
For visible calculation errors, sign errors, simplification errors, comparison errors, or wrong final expressions, include bbox around the exact handwritten expression that caused the deduction.
Only omit bbox when the deduction is conceptual or global and cannot be tied to a visible local area.
If the answer is blank or the expected answer is missing entirely, do not invent a bbox; explain it in reason.
similar_answer_review must be true when the student's answer appears mathematically valid but uses a different method, representation, counterexample, or reasoning from the model answer or rubric.
For valid similar answers, do not deduct solely because the method differs. Explain why in similar_answer_reason so a teacher can confirm it.
If the rubric has an [인정 유사답안] section and the student answer matches it, give appropriate credit and set similar_answer_review true.
Use the short step names from the rubric, for example "판단" or "반례제시", so the UI can show where points were lost.
If OCR uncertainty caused or may have caused a deduction, set ocr_related true and mention it in reason.
Use OCR normalized solution as primary evidence for diagrams, coordinate planes, graph shapes, plotted points, labels, and solution-step summaries that may not appear in OCR text or OCR LaTeX.
When the rubric awards graph or diagram points, do not deduct only because OCR text or LaTeX lacks that information if OCR normalized solution describes a matching graph/diagram with enough detail.
If OCR normalized solution says the graph/diagram is uncertain, give credit only for the visible features it clearly identifies and mark OCR-related deductions when needed.

{round_roles}

Max score: {max_score}
Question text: {question_text}
Model answer: {model_answer}
Rubric: {rubric_text}
Accepted similar answers:
{similar_answer_text or "None"}
OCR text: {ocr_text}
OCR normalized solution and visual notes: {ocr_normalized}
OCR LaTeX: {ocr_latex if formula_recognition_enabled else "Formula recognition was disabled for this exam."}
"""
        response = client.chat.completions.create(
            model=settings.openai_model,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        try:
            parsed = _parse_json_object(content)
        except ValueError:
            parsed = _repair_json_object_with_openai(client, content)
        reviews = parsed.get("reviews", [])[:4]
        fast_confidence = max((float(item.get("confidence", 0) or 0) for item in reviews), default=0)
        if fast and (ocr_needs_review or fast_confidence < 0.75):
            return await grade_submission_question(db, submission_id, question_id, fast=False)

    scores = []
    review_rows = []
    deduction_annotations = []
    final_review_index = len(reviews) - 1
    seen_annotation_keys = set()
    for review_index, review in enumerate(reviews):
        score = max(0, min(float(review.get("score", 0)), max_score))
        scores.append(score)
        deductions = review.get("deductions", [])
        deduction_text = json.dumps(deductions, ensure_ascii=False) if isinstance(deductions, list) else str(review.get("deduction_reasons", ""))
        if review_index == final_review_index and isinstance(deductions, list):
            for deduction in deductions:
                if isinstance(deduction, dict):
                    label = f"{deduction.get('step_name', '감점')} -{deduction.get('points_lost', '')}"
                    bbox = deduction.get("bbox") if isinstance(deduction.get("bbox"), dict) else {}
                    annotation_key = (
                        label,
                        str(deduction.get("reason", "")),
                        str(deduction.get("evidence_text", "")),
                        bbox.get("x"),
                        bbox.get("y"),
                        bbox.get("width"),
                        bbox.get("height"),
                    )
                    if annotation_key not in seen_annotation_keys:
                        seen_annotation_keys.add(annotation_key)
                        deduction_annotations.append(({**deduction, "label": label}, label))
        review_rows.append({
            "round_no": int(review.get("round_no", len(scores))),
            "score": score,
            "confidence": float(review.get("confidence", 0)),
            "reasoning": str(review.get("reasoning", "")),
            "deduction_reasons": deduction_text,
            "missing_steps": str(review.get("missing_steps", "")),
            "ocr_risk": str(review.get("ocr_risk", "unknown")),
            "similar_answer_review": bool(review.get("similar_answer_review", False)),
            "similar_answer_reason": str(review.get("similar_answer_reason", "")),
        })
    final_score = scores[-1] if scores else 0
    max_deviation = max(scores) - min(scores) if scores else 0
    review_reasons = []
    if ocr_needs_review and ocr_confidence < settings.ocr_confidence_threshold:
        review_reasons.append("ocr_review")
    if ocr_confidence < settings.ocr_confidence_threshold:
        review_reasons.append("ocr_low_confidence")
    if max_deviation >= max(2, max_score * 0.3):
        review_reasons.append("score_deviation")
    if not reviews or any(float(item.get("confidence", 0) or 0) < 0.3 for item in reviews):
        review_reasons.append("grading_uncertain")
    review_reasons = list(dict.fromkeys(review_reasons))
    needs_review = bool(review_reasons)
    final_reason = reviews[-1].get("reasoning", "") if reviews else "No review generated."
    similar_answer_review = bool(reviews[-1].get("similar_answer_review", False)) if reviews else False
    similar_answer_reason = str(reviews[-1].get("similar_answer_reason", "")) if reviews else ""
    with DB_WRITE_LOCK:
        db.query(models.GradingReview).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.query(models.AnswerAnnotation).filter_by(submission_id=submission_id, question_id=question_id, annotation_type="deduction").delete()
        for item, label in deduction_annotations:
            _add_annotation(db, submission_id, question_id, "deduction", item, label)
        for row in review_rows:
            db.add(models.GradingReview(submission_id=submission_id, question_id=question_id, **row))
        db.query(models.FinalScore).filter_by(submission_id=submission_id, question_id=question_id).delete()
        db.add(models.FinalScore(
            submission_id=submission_id,
            question_id=question_id,
            final_score=final_score,
            final_reason=final_reason,
            needs_teacher_review=needs_review,
        ))
        if not str(ocr_text or ocr_latex or ocr_normalized or "").strip():
            _add_annotation(db, submission_id, question_id, "blank", {
                "label": "무응답",
                "reason": "답안이 비어 있거나 판독 가능한 풀이가 없습니다.",
                "confidence": 1,
            }, "무응답")
        db.commit()
    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "final_score": final_score,
        "needs_teacher_review": needs_review,
        "review_scores": scores,
        "max_deviation": max_deviation,
        "final_reason": final_reason,
        "review_reasons": review_reasons,
        "similar_answer_review": similar_answer_review,
        "similar_answer_reason": similar_answer_reason,
    }


async def grade_submission_question_cached(db: Session, submission_id: int, question_id: int, force: bool = False) -> dict:
    existing = db.query(models.FinalScore).filter_by(submission_id=submission_id, question_id=question_id).first()
    reviews = db.query(models.GradingReview).filter_by(submission_id=submission_id, question_id=question_id).all()
    if existing and reviews and not force:
        scores = [review.score for review in reviews]
        max_deviation = max(scores) - min(scores) if scores else 0
        return {
            "submission_id": submission_id,
            "question_id": question_id,
            "final_score": existing.teacher_override_score if existing.teacher_override_score is not None else existing.final_score,
            "needs_teacher_review": existing.needs_teacher_review,
            "review_scores": scores,
            "max_deviation": max_deviation,
            "final_reason": existing.final_reason,
            "review_reasons": [],
            "similar_answer_review": any(review.similar_answer_review for review in reviews),
            "similar_answer_reason": next((review.similar_answer_reason for review in reversed(reviews) if review.similar_answer_reason), ""),
        }
    return await grade_submission_question(db, submission_id, question_id)


async def _process_grading_job(job_id: str) -> None:
    from app.database import SessionLocal

    job = GRADING_JOBS[job_id]
    db = SessionLocal()
    ocr_sem = asyncio.Semaphore(_sqlite_safe_concurrency(settings.ocr_concurrency))
    grade_sem = asyncio.Semaphore(_sqlite_safe_concurrency(settings.grading_concurrency))
    try:
        job["status"] = "running"
        exam_id = job["exam_id"]
        force = job["force"]
        submissions = db.query(models.Submission).filter_by(exam_id=exam_id).all()
        questions = db.query(models.Question).filter_by(exam_id=exam_id).order_by(models.Question.question_no).all()
        submission_jobs = [(item.id, f"{item.student.student_no} {item.student.name}") for item in submissions]
        question_jobs = [(item.id, item.question_no) for item in questions]
        db.commit()
        job["total"] = len(submission_jobs) * len(question_jobs)
        job["done"] = 0
        failed_ocr_keys: set[tuple[int, int]] = set()

        async def run_submission_ocr(submission_id: int, student_label: str) -> None:
            async with ocr_sem:
                local_db = SessionLocal()
                try:
                    job["label"] = f"{student_label} OCR 중"
                    await create_ocr_for_submission(local_db, submission_id, force=force)
                    ocr_results = local_db.query(models.OcrResult).filter_by(submission_id=submission_id).all()
                    for ocr in ocr_results:
                        normalized = ocr.normalized_result or ""
                        if ocr.needs_review and "AI image interpretation failed" in normalized:
                            failed_ocr_keys.add((submission_id, ocr.question_id))
                            question_no = next((no for qid, no in question_jobs if qid == ocr.question_id), ocr.question_id)
                            job["failed_items"].append({
                                "submission_id": submission_id,
                                "question_id": ocr.question_id,
                                "student_label": student_label,
                                "question_label": f"{question_no}번",
                                "stage": "ocr",
                                "error": normalized.splitlines()[1] if len(normalized.splitlines()) > 1 else "AI image interpretation failed",
                            })
                except Exception as exc:  # noqa: BLE001
                    for question_id, question_no in question_jobs:
                        failed_ocr_keys.add((submission_id, question_id))
                        job["failed_items"].append({
                            "submission_id": submission_id,
                            "question_id": question_id,
                            "student_label": student_label,
                            "question_label": f"{question_no}번",
                            "stage": "ocr",
                            "error": f"{type(exc).__name__}: {exc}",
                        })
                finally:
                    local_db.close()

        await asyncio.gather(*(run_submission_ocr(submission_id, student_label) for submission_id, student_label in submission_jobs))

        async def run_grade(submission_id: int, student_label: str, question_id: int, question_no: int) -> None:
            async with grade_sem:
                local_db = SessionLocal()
                try:
                    job["label"] = f"{student_label} / {question_no}번 채점 중"
                    if (submission_id, question_id) in failed_ocr_keys:
                        result = _fallback_grade_for_failed_ocr(local_db, submission_id, question_id)
                    else:
                        result = await grade_submission_question_cached(local_db, submission_id, question_id, force=force)
                    if result.get("needs_teacher_review"):
                        job["review_needed_count"] += 1
                except Exception as exc:  # noqa: BLE001
                    try:
                        local_db.rollback()
                        fallback = _fallback_grade_for_failed_ai(
                            local_db,
                            submission_id,
                            question_id,
                            f"AI 채점에 실패하여 교사 확인이 필요합니다: {type(exc).__name__}: {exc}",
                        )
                        if fallback.get("needs_teacher_review"):
                            job["review_needed_count"] += 1
                    except Exception as fallback_exc:  # noqa: BLE001
                        local_db.rollback()
                        job["failed_items"].append({
                            "submission_id": submission_id,
                            "question_id": question_id,
                            "student_label": student_label,
                            "question_label": f"{question_no}번",
                            "stage": "fallback",
                            "error": f"Fallback save failed: {type(fallback_exc).__name__}: {fallback_exc}",
                        })
                    job["failed_items"].append({
                        "submission_id": submission_id,
                        "question_id": question_id,
                        "student_label": student_label,
                        "question_label": f"{question_no}번",
                        "stage": "grade",
                        "error": f"{type(exc).__name__}: {exc}",
                    })
                finally:
                    job["done"] += 1
                    local_db.close()

        await asyncio.gather(*(
            run_grade(submission_id, student_label, question_id, question_no)
            for submission_id, student_label in submission_jobs
            for question_id, question_no in question_jobs
        ))
        job["status"] = "done" if not job["failed_items"] else "failed"
        job["label"] = "전체 채점 완료"
    except Exception as exc:  # noqa: BLE001
        job["status"] = "failed"
        job["label"] = f"작업 실패: {type(exc).__name__}: {exc}"
    finally:
        job["finished_at"] = time.time()
        db.close()


def _run_grading_job_thread(job_id: str) -> None:
    asyncio.run(_process_grading_job(job_id))


def start_grading_job(exam_id: int, force: bool = False) -> dict:
    job_id = uuid.uuid4().hex
    GRADING_JOBS[job_id] = {
        "job_id": job_id,
        "exam_id": exam_id,
        "status": "pending",
        "done": 0,
        "total": 0,
        "label": "채점 작업 대기 중",
        "force": force,
        "failed_items": [],
        "review_needed_count": 0,
        "created_at": time.time(),
    }
    threading.Thread(target=_run_grading_job_thread, args=(job_id,), daemon=True).start()
    return GRADING_JOBS[job_id]


def get_grading_job(job_id: str) -> dict:
    job = GRADING_JOBS.get(job_id)
    if not job:
        raise ValueError("Grading job not found")
    return job


def _submission_ids_for_exam(db: Session, exam_id: int) -> list[int]:
    return [item[0] for item in db.query(models.Submission.id).filter_by(exam_id=exam_id).all()]


def clear_exam_submissions(db: Session, exam_id: int) -> dict:
    submission_ids = _submission_ids_for_exam(db, exam_id)
    with DB_WRITE_LOCK:
        deleted = {"annotations": 0, "ocr_results": 0, "grading_reviews": 0, "final_scores": 0, "submissions": 0}
        if submission_ids:
            deleted["annotations"] = db.query(models.AnswerAnnotation).filter(models.AnswerAnnotation.submission_id.in_(submission_ids)).delete(synchronize_session=False)
            deleted["ocr_results"] = db.query(models.OcrResult).filter(models.OcrResult.submission_id.in_(submission_ids)).delete(synchronize_session=False)
            deleted["grading_reviews"] = db.query(models.GradingReview).filter(models.GradingReview.submission_id.in_(submission_ids)).delete(synchronize_session=False)
            deleted["final_scores"] = db.query(models.FinalScore).filter(models.FinalScore.submission_id.in_(submission_ids)).delete(synchronize_session=False)
        deleted["submissions"] = db.query(models.Submission).filter_by(exam_id=exam_id).delete(synchronize_session=False)
        db.commit()
    exam_dir = Path(settings.processed_dir) / f"exam_{exam_id}"
    shutil.rmtree(exam_dir / "submissions", ignore_errors=True)
    shutil.rmtree(exam_dir / "ocr", ignore_errors=True)
    shutil.rmtree(exam_dir / "region_previews", ignore_errors=True)
    shutil.rmtree(exam_dir / "previews", ignore_errors=True)
    return deleted


def clear_exam_students_and_submissions(db: Session, exam_id: int) -> dict:
    deleted = clear_exam_submissions(db, exam_id)
    with DB_WRITE_LOCK:
        deleted["students"] = db.query(models.Student).filter_by(exam_id=exam_id).delete(synchronize_session=False)
        db.commit()
    return deleted


def delete_grading_results(db: Session, exam_id: int, question_id: int | None = None) -> dict:
    submission_ids = _submission_ids_for_exam(db, exam_id)
    if not submission_ids:
        return {"deleted": True, "annotations": 0, "ocr_results": 0, "grading_reviews": 0, "final_scores": 0}
    with DB_WRITE_LOCK:
        filters = [models.AnswerAnnotation.submission_id.in_(submission_ids)]
        if question_id is not None:
            filters.append(models.AnswerAnnotation.question_id == question_id)
        annotations = db.query(models.AnswerAnnotation).filter(*filters).delete(synchronize_session=False)

        filters = [models.OcrResult.submission_id.in_(submission_ids)]
        if question_id is not None:
            filters.append(models.OcrResult.question_id == question_id)
        ocr_results = db.query(models.OcrResult).filter(*filters).delete(synchronize_session=False)

        filters = [models.GradingReview.submission_id.in_(submission_ids)]
        if question_id is not None:
            filters.append(models.GradingReview.question_id == question_id)
        grading_reviews = db.query(models.GradingReview).filter(*filters).delete(synchronize_session=False)

        filters = [models.FinalScore.submission_id.in_(submission_ids)]
        if question_id is not None:
            filters.append(models.FinalScore.question_id == question_id)
        final_scores = db.query(models.FinalScore).filter(*filters).delete(synchronize_session=False)
        db.commit()

    exam_dir = Path(settings.processed_dir) / f"exam_{exam_id}"
    if question_id is None:
        shutil.rmtree(exam_dir / "ocr", ignore_errors=True)
        shutil.rmtree(exam_dir / "region_previews", ignore_errors=True)
    else:
        for path in (exam_dir / "ocr").glob(f"submission_*/question_{question_id}.png"):
            path.unlink(missing_ok=True)
        for path in (exam_dir / "region_previews").glob(f"submission_*/question_{question_id}.png"):
            path.unlink(missing_ok=True)
    return {
        "deleted": True,
        "annotations": annotations,
        "ocr_results": ocr_results,
        "grading_reviews": grading_reviews,
        "final_scores": final_scores,
    }


async def _process_question_regrade_job(job_id: str, question_id: int) -> None:
    from app.database import SessionLocal

    job = GRADING_JOBS[job_id]
    db = SessionLocal()
    try:
        job["status"] = "running"
        exam_id = job["exam_id"]
        delete_grading_results(db, exam_id, question_id)
        submissions = db.query(models.Submission).filter_by(exam_id=exam_id).all()
        question = db.get(models.Question, question_id)
        if not question:
            raise ValueError("Question not found")
        submission_jobs = [(item.id, f"{item.student.student_no} {item.student.name}") for item in submissions]
        question_no = question.question_no
        db.commit()
        job["total"] = len(submission_jobs)
        job["done"] = 0
        sem = asyncio.Semaphore(_sqlite_safe_concurrency(settings.grading_concurrency))

        async def run_one(submission_id: int, student_label: str) -> None:
            async with sem:
                local_db = SessionLocal()
                try:
                    job["label"] = f"{student_label} / {question_no}번 재채점 중"
                    await create_ocr_for_submission_question(local_db, submission_id, question_id, force=True)
                    result = await grade_submission_question_cached(local_db, submission_id, question_id, force=True)
                    if result.get("needs_teacher_review"):
                        job["review_needed_count"] += 1
                except Exception as exc:  # noqa: BLE001
                    try:
                        local_db.rollback()
                        fallback = _fallback_grade_for_failed_ai(
                            local_db,
                            submission_id,
                            question_id,
                            f"AI 재채점에 실패하여 교사 확인이 필요합니다: {type(exc).__name__}: {exc}",
                        )
                        if fallback.get("needs_teacher_review"):
                            job["review_needed_count"] += 1
                    except Exception as fallback_exc:  # noqa: BLE001
                        local_db.rollback()
                        job["failed_items"].append({
                            "submission_id": submission_id,
                            "question_id": question_id,
                            "student_label": student_label,
                            "question_label": f"{question_no}번",
                            "stage": "fallback",
                            "error": f"Fallback save failed: {type(fallback_exc).__name__}: {fallback_exc}",
                        })
                    job["failed_items"].append({
                        "submission_id": submission_id,
                        "question_id": question_id,
                        "student_label": student_label,
                        "question_label": f"{question_no}번",
                        "stage": "grade",
                        "error": f"{type(exc).__name__}: {exc}",
                    })
                finally:
                    job["done"] += 1
                    local_db.close()

        await asyncio.gather(*(run_one(submission_id, student_label) for submission_id, student_label in submission_jobs))
        job["status"] = "done" if not job["failed_items"] else "failed"
        job["label"] = "문항별 재채점 완료"
    except Exception as exc:  # noqa: BLE001
        job["status"] = "failed"
        job["label"] = f"작업 실패: {type(exc).__name__}: {exc}"
    finally:
        job["finished_at"] = time.time()
        db.close()


def start_question_regrade_job(exam_id: int, question_id: int) -> dict:
    job_id = uuid.uuid4().hex
    GRADING_JOBS[job_id] = {
        "job_id": job_id,
        "exam_id": exam_id,
        "status": "pending",
        "done": 0,
        "total": 0,
        "label": "문항별 재채점 대기 중",
        "force": True,
        "failed_items": [],
        "review_needed_count": 0,
        "created_at": time.time(),
    }
    threading.Thread(target=lambda: asyncio.run(_process_question_regrade_job(job_id, question_id)), daemon=True).start()
    return GRADING_JOBS[job_id]


def save_teacher_override(
    db: Session,
    submission_id: int,
    question_id: int,
    score: float,
    note: str = "",
    clear_review: bool = True,
) -> dict:
    question = db.get(models.Question, question_id)
    if not question:
        raise ValueError("Question not found")
    final = db.query(models.FinalScore).filter_by(submission_id=submission_id, question_id=question_id).first()
    if not final:
        final = models.FinalScore(submission_id=submission_id, question_id=question_id, final_score=0)
        db.add(final)
    final.teacher_override_score = max(0, min(score, question.max_score))
    final.teacher_note = note
    final.teacher_reviewed = True
    final.teacher_reviewed_at = datetime.utcnow()
    if clear_review:
        final.needs_teacher_review = False
    db.commit()
    db.refresh(final)
    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "final_score": final.final_score,
        "teacher_override_score": final.teacher_override_score,
        "effective_score": final.teacher_override_score if final.teacher_override_score is not None else final.final_score,
        "teacher_note": final.teacher_note,
        "teacher_reviewed": final.teacher_reviewed,
    }


def add_rubric_note(db: Session, question_id: int, section: str, text_value: str) -> models.Question:
    question = db.get(models.Question, question_id)
    if not question:
        raise ValueError("Question not found")
    clean_section = section.strip() or "인정 유사답안"
    clean_text = text_value.strip()
    if not clean_text:
        raise ValueError("Rubric note text is required")
    marker = f"[{clean_section}]"
    rubric = question.rubric_text or ""
    note_line = f"- {clean_text}"
    if marker in rubric:
        question.rubric_text = rubric.rstrip() + f"\n{note_line}"
    else:
        question.rubric_text = rubric.rstrip() + f"\n\n{marker}\n{note_line}"
    db.commit()
    db.refresh(question)
    return question


def create_similar_answer(db: Session, question_id: int, title: str, text_value: str, source_reason: str = "") -> models.SimilarAnswer:
    question = db.get(models.Question, question_id)
    if not question:
        raise ValueError("Question not found")
    clean_text = text_value.strip()
    if not clean_text:
        raise ValueError("Similar answer text is required")
    similar_answer = models.SimilarAnswer(
        question_id=question_id,
        title=title.strip() or "유사답안",
        text=clean_text,
        source_reason=source_reason.strip(),
    )
    db.add(similar_answer)
    db.commit()
    db.refresh(similar_answer)
    return similar_answer


def update_similar_answer(db: Session, answer_id: int, title: str, text_value: str, source_reason: str = "") -> models.SimilarAnswer:
    similar_answer = db.get(models.SimilarAnswer, answer_id)
    if not similar_answer:
        raise ValueError("Similar answer not found")
    clean_text = text_value.strip()
    if not clean_text:
        raise ValueError("Similar answer text is required")
    similar_answer.title = title.strip() or "유사답안"
    similar_answer.text = clean_text
    similar_answer.source_reason = source_reason.strip()
    similar_answer.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(similar_answer)
    return similar_answer


def delete_similar_answer(db: Session, answer_id: int) -> None:
    similar_answer = db.get(models.SimilarAnswer, answer_id)
    if not similar_answer:
        raise ValueError("Similar answer not found")
    db.delete(similar_answer)
    db.commit()


def generate_similar_answer_rubric(db: Session, question_id: int, similar_answer_reason: str) -> str:
    question = db.get(models.Question, question_id)
    if not question:
        raise ValueError("Question not found")
    reason = similar_answer_reason.strip()
    if not reason:
        raise ValueError("Similar answer reason is required")
    if not settings.openai_api_key:
        return "\n".join([
            "유사답안: [제목 확인 필요]",
            f"1. 접근확인 ({round(question.max_score * 0.3, 1):g}점): 기존 풀이와 다른 정당한 접근을 사용 [확인 필요]",
            f"2. 조건충족 ({round(question.max_score * 0.3, 1):g}점): 문제 조건을 만족함 [확인 필요]",
            f"3. 결론일치 ({round(question.max_score * 0.4, 1):g}점): 모범답안과 동등한 결론을 도출 [확인 필요]",
        ])
    prompt = f"""
Create a concise Korean step-by-step rubric for an accepted similar answer.
Return only plain text, no JSON, no markdown fences.

Format exactly:
유사답안: short title
1. step name (points점): criterion
2. step name (points점): criterion
3. step name (points점): criterion

Rules:
- Total points should equal {question.max_score} when possible.
- Use the student's similar-answer reason to make concrete criteria.
- Keep step names short.
- Do not add general cautions.

Question: {question.question_text}
Model answer: {question.model_answer}
Existing rubric: {question.rubric_text}
Similar answer reason: {reason}
"""
    client = get_openai_client()
    response = client.chat.completions.create(
        model=settings.openai_model,
        messages=[{"role": "user", "content": prompt}],
    )
    return (response.choices[0].message.content or "").strip()


def export_scores_excel(db: Session, exam_id: int) -> Path:
    exam = db.get(models.Exam, exam_id)
    questions = db.query(models.Question).filter_by(exam_id=exam_id).order_by(models.Question.question_no).all()
    submissions = db.query(models.Submission).filter_by(exam_id=exam_id).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "student_scores"
    headers = ["학번", "이름"] + [f"{q.question_no}번" for q in questions] + ["총점", "검토필요문항"]
    sheet.append(headers)
    for submission in submissions:
        row = [submission.student.student_no, submission.student.name]
        total = 0.0
        warnings = []
        for question in questions:
            final = db.query(models.FinalScore).filter_by(submission_id=submission.id, question_id=question.id).first()
            score = final.teacher_override_score if final and final.teacher_override_score is not None else final.final_score if final else 0
            row.append(score)
            total += float(score)
            if final and final.needs_teacher_review:
                warnings.append(f"{question.question_no}번")
        row.extend([total, ", ".join(warnings)])
        sheet.append(row)
    out_path = Path(settings.export_dir) / f"exam_{exam.id}_scores.xlsx"
    workbook.save(out_path)
    return out_path


def get_grading_results(db: Session, exam_id: int) -> list[dict]:
    questions = db.query(models.Question).filter_by(exam_id=exam_id).order_by(models.Question.question_no).all()
    submissions = db.query(models.Submission).filter_by(exam_id=exam_id).all()
    rows = []
    for submission in submissions:
        question_results = []
        for question in questions:
            final = db.query(models.FinalScore).filter_by(submission_id=submission.id, question_id=question.id).first()
            reviews = db.query(models.GradingReview).filter_by(
                submission_id=submission.id,
                question_id=question.id,
            ).order_by(models.GradingReview.round_no).all()
            ocr = db.query(models.OcrResult).filter_by(submission_id=submission.id, question_id=question.id).first()
            final_review = reviews[-1] if reviews else None
            review_scores = [review.score for review in reviews]
            max_deviation = max(review_scores) - min(review_scores) if review_scores else 0
            review_reasons = _review_reasons(question, ocr, review_scores, [
                {"confidence": review.confidence} for review in reviews
            ]) if ocr else []
            deductions = _parse_deductions(final_review.deduction_reasons if final_review else "")
            annotations = db.query(models.AnswerAnnotation).filter_by(
                submission_id=submission.id,
                question_id=question.id,
            ).all()
            effective_score = final.teacher_override_score if final and final.teacher_override_score is not None else final.final_score if final else None
            question_results.append({
                "question_id": question.id,
                "question_no": question.question_no,
                "max_score": question.max_score,
                "final_score": effective_score,
                "ai_score": final.final_score if final else None,
                "teacher_override_score": final.teacher_override_score if final else None,
                "teacher_note": final.teacher_note if final else "",
                "teacher_reviewed": bool(final.teacher_reviewed) if final else False,
                "needs_teacher_review": bool(final.needs_teacher_review) if final else False,
                "final_reason": final.final_reason if final else "",
                "deduction_reasons": final_review.deduction_reasons if final_review else "",
                "missing_steps": final_review.missing_steps if final_review else "",
                "ocr_risk": final_review.ocr_risk if final_review else "",
                "ocr_confidence": ocr.confidence if ocr else None,
                "ocr_text": ocr.text_result if ocr else "",
                "ocr_needs_review": bool(ocr.needs_review) if ocr else False,
                "max_deviation": max_deviation,
                "review_reasons": review_reasons,
                "deductions": deductions,
                "similar_answer_review": bool(final_review.similar_answer_review) if final_review else False,
                "similar_answer_reason": final_review.similar_answer_reason if final_review else "",
                "annotations": [{
                    "annotation_type": annotation.annotation_type,
                    "label": annotation.label,
                    "reason": annotation.reason,
                    "evidence_text": annotation.evidence_text,
                    "x": annotation.x,
                    "y": annotation.y,
                    "width": annotation.width,
                    "height": annotation.height,
                    "confidence": annotation.confidence,
                } for annotation in annotations],
                "reviews": [{
                    "round_no": review.round_no,
                    "score": review.score,
                    "confidence": review.confidence,
                    "reasoning": review.reasoning,
                    "deduction_reasons": review.deduction_reasons,
                    "missing_steps": review.missing_steps,
                    "ocr_risk": review.ocr_risk,
                } for review in reviews],
            })
        rows.append({
            "submission_id": submission.id,
            "student_no": submission.student.student_no,
            "student_name": submission.student.name,
            "questions": question_results,
            "total_score": sum(float(item["final_score"] or 0) for item in question_results),
            "needs_teacher_review_count": sum(1 for item in question_results if item["needs_teacher_review"] or item["ocr_needs_review"]),
        })
    return rows


def delete_exam(db: Session, exam_id: int) -> None:
    exam = db.get(models.Exam, exam_id)
    if not exam:
        raise ValueError("Exam not found")

    submission_ids = [item[0] for item in db.query(models.Submission.id).filter_by(exam_id=exam_id).all()]
    question_ids = [item[0] for item in db.query(models.Question.id).filter_by(exam_id=exam_id).all()]

    if submission_ids:
        db.query(models.AnswerAnnotation).filter(models.AnswerAnnotation.submission_id.in_(submission_ids)).delete(synchronize_session=False)
        db.query(models.FinalScore).filter(models.FinalScore.submission_id.in_(submission_ids)).delete(synchronize_session=False)
        db.query(models.GradingReview).filter(models.GradingReview.submission_id.in_(submission_ids)).delete(synchronize_session=False)
        db.query(models.OcrResult).filter(models.OcrResult.submission_id.in_(submission_ids)).delete(synchronize_session=False)
    if question_ids:
        db.query(models.SimilarAnswer).filter(models.SimilarAnswer.question_id.in_(question_ids)).delete(synchronize_session=False)
        db.query(models.Region).filter(models.Region.question_id.in_(question_ids)).delete(synchronize_session=False)

    db.query(models.Region).filter_by(exam_id=exam_id).delete(synchronize_session=False)
    db.query(models.Submission).filter_by(exam_id=exam_id).delete(synchronize_session=False)
    db.query(models.Student).filter_by(exam_id=exam_id).delete(synchronize_session=False)
    db.query(models.Question).filter_by(exam_id=exam_id).delete(synchronize_session=False)
    db.delete(exam)
    db.commit()

    shutil.rmtree(Path(settings.processed_dir) / f"exam_{exam_id}", ignore_errors=True)
    for upload_path in Path(settings.upload_dir).glob(f"exam_{exam_id}_*"):
        if upload_path.is_file():
            upload_path.unlink(missing_ok=True)
        elif upload_path.is_dir():
            shutil.rmtree(upload_path, ignore_errors=True)

    export_path = Path(settings.export_dir) / f"exam_{exam_id}_scores.xlsx"
    export_path.unlink(missing_ok=True)
